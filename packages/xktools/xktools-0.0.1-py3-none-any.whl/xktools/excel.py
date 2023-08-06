# -*- coding: utf-8 -*-
# @Time    : 2020/3/23 8:49
# @Author  : david
# @Email   : davidxiaocn@qq.com

import sys  # 文件处理
import time  # 时间处理
from datetime import datetime

import xlrd  # xlrd 用于读取数据，速度快
from openpyxl import load_workbook  # openpyxl基于模版写文件
from xlrd import xldate_as_tuple

"""
本模块需要外部依赖的库
pip install xlrd           读数据，加快速度
pip install openpyxl       读数据和写数据，主要用于文件保存
pip install pymysql        MySQL数据库支持；如果只是本地SQLite不需要
生成库 xktools ，导入之后可以执行 
"""
"""
实现目标：
    自动对Excel表格进行数据库导入和导出，数据加工、Excel汇总、分发，主要实现自动对应字段。
实现思路：
    为了便于处理数据，把Excel分为基础表和汇总表。通过数据库对基础表进行处理，汇总通过Excel自身功能实现，简化对Excel的处理。
    基础表，格式为第一行数据类型（文本、数字、大文本），第二行为字段名称【可以自定义，需要设置字段类型和字段名称】。
    数据汇总，基于基础表采用Excel的汇总功能，通过模版提前设置好数据汇总，只更新基础表即可，也可以在数据库中汇总好，直接导出结果。
    根据Excel设置，自动生成创建表、插入数据、读取数据的SQL，导入数据库SQLite（MySQL），脚本自动根据字段所在的位置定位。
    到如数据库之后，在数据库中存放，加工，可以根据模版导出到Excel，不受字段所在行列限制。
    可以实现对Excel汇总合并、分发拆分。

不同人之间，确定明确的对应格式，不论是如何更新表格式，只要字段列名称不变，就可以进行对应

为了简化思路，封装对象为简单的函数调用
Excel之间只传递基础表
其他的表，采用自动汇总的模式。为了简化，采用SQLite进行数据交换
"""

'''
的对象格式
KMTools
    KMWorkBook【save_to_db();】
        KMSheet【save_to_db()】
            KMSheetSQL
'''


def excel_save_to_db(file_name, sheet_name, db, table_name, is_create_table=0, is_delete_data=1, is_print_progress=0):
    """
    电子表格导入到MYSQL数据库，文件名、表单名称、要导入的数据库表名称
    是否新创建，是否删除数据，是否打印导入过程
    """
    wb = KmWorkBook(file_name)
    wb.save_to_db(sheet_name, db, table_name, is_create_table, is_delete_data, is_print_progress)


def db_read_to_sheet(db, table_name, file_name, sheet_name, save_file_name, sql_where=""):
    # MYSQL数据导出到电子表格,本函只能能导入一个sheet
    # 如果导出多个sheet，需要调用对象，然后执行多次 read_to_sheet
    wb = KmWorkBook(file_name)
    wb.read_to_sheet(db, table_name, sheet_name, sql_where)
    wb.save_file(save_file_name)
    return wb


class KmWorkSheetSQL(object):
    """
    根据Excel 第一行字段格式和第二行字段名称获取字段信息，自动成圣创建、 删除、插入等SQL
    获取字段列表的数据类型：
    1 字符串 2 数字  3长文本，为了简化，只做3中数据格式设置
    数据格式、日期格式等显示通过电子表格的格式设置
    """

    def __init__(self, table_ame):
        self.db_Kind = "MySQL"
        self.table_ame = table_ame
        self.is_star = "yes"

        self.sql_drop_table = ""  # 删除表语句 DROP Table 语句
        self.sql_delete_table = ""  # 删除数据语句 DELETE Table 语句
        self.sql_create_table = ""  # 创建表 CREATE Table 语句
        self.sql_create_table_filed = ""  # 创建表字段 CREATE Table 语句
        self.sql_select_table = ""  # Select 语句
        self.sql_select_table_field = ""  # Select 语句 字段
        self.field_columns = []  # 字段所在的EXCEL列

    def db_kind(self, kind):
        self.db_Kind = kind

    def add_field(self, field_kind, cell_field, column):
        # 如果没有设置或者为空，则忽略
        if field_kind != '' and field_kind is not None:
            # 判读第一个字符是否为数字
            if str.isdigit(str(field_kind)[0]) :
                field_kind = int(field_kind)
            else:
                field_kind = 1
                print('字段格式设置错误，请输入1、2、3分别代表字符、数字、大文本！，已经按照文本方式设置')
        if field_kind == 1 or field_kind == 2 or field_kind == 3:
            cell_field = cell_field.replace('-', '')
            cell_field = cell_field.replace('\n', '')
            cell_field = cell_field.replace('（', '')
            cell_field = cell_field.replace(')', '')
            cell_field = cell_field.replace('（', '')
            cell_field = cell_field.replace('）', '')
            # 以上去除非法字符
            #if is_number(cell_field[0]):
            if str.isdigit(cell_field[0]):
                cell_field = "A" + cell_field
            if self.is_star != "yes":
                self.sql_select_table_field = self.sql_select_table_field + "," + cell_field
            else:
                self.sql_select_table_field = cell_field
            self.is_star = "no"
            self.field_columns.append(column)

        if field_kind == 1:
            self.sql_create_table_filed = self.sql_create_table_filed + ", `" + cell_field + "` varchar(100)"
        elif field_kind == 2:
            self.sql_create_table_filed = self.sql_create_table_filed + ", `" + cell_field + "` float"
        elif field_kind == 3:
            # text
            self.sql_create_table_filed = self.sql_create_table_filed + ", `" + cell_field + "`  text "

        # SQLite数据库 SQL 语句
        if self.db_Kind == "SQLite":
            self.sql_drop_table = "DROP TABLE   IF   EXISTS	`" + self.table_ame + "`"
            self.sql_delete_table = "DELETE FROM `" + self.table_ame + "`"
            self.sql_select_table = "SELECT  " + self.sql_select_table_field + "  FROM  " + self.table_ame
            self.sql_create_table = "CREATE TABLE IF NOT EXISTS	`" + self.table_ame + "` (\
                        `ID` INTEGER PRIMARY KEY   AUTOINCREMENT  , \
                        `Entity_NUM` varchar(20)" \
                                    + self.sql_create_table_filed + "       )  "
        # MySQL数据库 SQL 语句
        if self.db_Kind == "MySQL":
            self.sql_drop_table = "DROP TABLE   IF   EXISTS	`" + self.table_ame + "`"
            self.sql_delete_table = "DELETE FROM `" + self.table_ame + "`"
            self.sql_select_table = "SELECT  " + self.sql_select_table_field + "  FROM  " + self.table_ame
            self.sql_create_table = "CREATE TABLE IF NOT EXISTS	`" + self.table_ame + "` (\
                        `ID` int(11) NOT NULL AUTO_INCREMENT, \
                        `Entity_NUM` varchar(20) " \
                                    + self.sql_create_table_filed + "  ,  PRIMARY KEY (`ID`)  \
                        ) ENGINE = MyISAM DEFAULT CHARSET = utf8 AUTO_INCREMENT = 1 "


def get_time_stamp():
    return "-" + time.strftime("%Y%m%d", time.localtime())


class KmWorkBook:
    """
    kmWorkBook,打开的文件对象，可以是xlrd，只读，或者 openpyxl 写文件
    file_name， 文件名称
    """

    def __init__(self, file_name):
        self.file_name = file_name
        self.kmWorkBook = None
        self.sheet = None

    def open_file(self, file_name):
        self.file_name = file_name
        self.kmWorkBook = load_workbook(self.file_name)
        return self.kmWorkBook

    def sheets(self, sheet_name):
        return self.get_sheet(sheet_name)

    def get_sheet(self, sheet_name):
        return self.kmWorkBook.sheet_by_name(sheet_name)

    def save_to_db(self, sheet_name, db, table_name, is_create_table=0, is_delete_data=1, is_print_progress=0):
        # 为了加快读取数据的速度，采用xlrd模块进行读取
        self.kmWorkBook = xlrd.open_workbook(self.file_name)
        self.sheet = KmWorksheet(db, table_name, self.kmWorkBook.sheet_by_name(sheet_name), "read")
        self.sheet.save_to_db(is_create_table, is_delete_data, is_print_progress)

    def read_to_sheet(self, db, table_name, sheet_name, sql_where="", delete_col1=1):
        if self.kmWorkBook is None:
            self.kmWorkBook = load_workbook(self.file_name)
        sheet = KmWorksheet(db, table_name, self.kmWorkBook[sheet_name], "write")
        sheet.read_to_sheet(sql_where, delete_col1)



    def save_file(self, save_file_name):
        self.kmWorkBook.save(save_file_name)
        print("成功保存文件%s" % save_file_name)
        return save_file_name

    def get_worksheet_sql1(self, db, table_name, sheet_name):
        self.kmWorkBook = xlrd.open_workbook(self.file_name)
        sheet = KmWorksheet(db, table_name, self.kmWorkBook.sheet_by_name(sheet_name), "read")
        return sheet.kmWorkSheetSQL


class KmWorksheet(object):
    """
    sheet对象处理，根据模版对数据进行处理
    如果只是只读到数据库，采用 xlrd，以加快读入速度
    如果是写电子表格，需要用到 openpyxl，用于处理文件另存、格式化等工作
    写入数据，建议基于模版，提前设置好模版格式，直接写入数据
    """

    def __init__(self, db, table_name, sheet, read):
        self.is_create_table = 0
        self.is_delete_data = 1
        self.is_is_print_progress = 0

        self.kmWorkSheetSQL = None
        self.read_only = "write"
        self.sheet = sheet
        self.db = db

        self.table_name = table_name
        self.kmWorkSheetSQL = KmWorkSheetSQL(table_name)
        self.kmWorkSheetSQL.db_kind(self.db.db_kind)

        # 只读，加快速度
        if read == "read":
            print("最大列数：%d" % self.sheet.ncols)
            for i in range(0, self.sheet.ncols):
                field_kind = self.sheet.cell_value(0, i)
                cell_field = self.sheet.cell_value(1, i)
                self.kmWorkSheetSQL.add_field(field_kind, cell_field, i)
        else:
            print("最大列数：%d" % self.sheet.max_column)
            for i in range(1, self.sheet.max_column + 1):
                # 从1开始，最大列增加1
                field_kind = self.sheet.cell(row=1, column=i).value
                cell_field = self.sheet.cell(row=2, column=i).value
                if cell_field != "None":
                    # 处理标记了，但是实际为空格的数据
                    self.kmWorkSheetSQL.add_field(field_kind, cell_field, i)

    def save_to_db(self, is_create_table=0, is_delete_data=1, is_print_progress=0):
        # 把按照标准格式的数据导入到 SQLite 中
        if is_create_table == 1:
            self.db.execute(self.kmWorkSheetSQL.sql_drop_table)
            self.db.execute(self.kmWorkSheetSQL.sql_create_table)
        if is_delete_data == 1:
            # print("is_delete_data data = "+str(is_delete_data))
            self.db.execute(self.kmWorkSheetSQL.sql_delete_table)

        try:
            # 从第三行开始导入,0第一行为标记位；1第二行为字段名称；2第三行为数据
            for row in range(2, self.sheet.nrows):
                value = ""
                is_star = "yes"
                for field_col in self.kmWorkSheetSQL.field_columns:
                    cell_ctype = self.sheet.cell(row, field_col).ctype  # 获取单元格返回的数据类型
                    # print(cell_ctype)
                    ctype = int(self.sheet.cell(0, field_col).value)
                    cell_value = self.sheet.cell(row, field_col).value  # 获取单元格内容
                    if ctype == 2:                              # 有些版本数字需要处理为空的情况
                        if cell_value == "":
                            cell_value = "0"
                        if str.isdigit(str(cell_value)[0]):     # 判断子一个字母是否是数字
                            pass
                        else:
                            cell_value = "0"
                    if cell_ctype == 3:  # 如果是日期格式，需要转换一下
                        if cell_value == "0":  # 为0，需要特殊处理
                            cell_value = cell_value
                        else:
                            try:
                                date = datetime(*xldate_as_tuple(cell_value, 0))
                                cell_value = date.strftime('%Y/%m/%d')
                            except BaseException as e:  # 数值转日期格式会出错，特殊处理
                                print(e)
                                cell_value = cell_value
                    cell_value = str(cell_value)
                    if is_star != "yes":
                        value = value + "," + "'" + cell_value + "'"
                    else:
                        value = value + "'" + cell_value + "'"
                    is_star = "no"
                sql = "insert into " + self.table_name + "(" + self.kmWorkSheetSQL.sql_select_table_field \
                      + ")values(" + value + ")"
                # print(sql)  #打印导入数据的每条SQL
                if is_print_progress == 0:
                    print("导入第%d条数据" % row)
                self.db.execute(sql)
            sql = "update " + self.table_name + "  set Entity_NUM = ID"
            self.db.execute(sql)
        except BaseException as e:
            print(e)
            print("ERR:insert =  " + self.table_name)
            info = sys.exc_info()
            print(info[0], ":", info[1])
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))

    def read_to_sheet(self, sql_where="", delete_col1=1):
        # col 生成的SQL中字段所在的列  和数据库返回的数据所在的列对应
        # 这是根据Excel字段所在列和字段对应的关键
        # 推荐根据模版生成文件，所以采用 openpyxl 生成文件，因为需要根据设置的列赋值，所以按照cell
        data = self.db.fetchall(self.kmWorkSheetSQL.sql_select_table + " " + sql_where)
        for row, rowData in enumerate(data):
            for col, value in enumerate(rowData):
                # 返回格式说明： cell(row, column, value=None)，return cell
                self.sheet.cell(row=row + 3, column=self.kmWorkSheetSQL.field_columns[col]).value = value
        if delete_col1 != 1:
            self.sheet.delete_rows(1, 1)

