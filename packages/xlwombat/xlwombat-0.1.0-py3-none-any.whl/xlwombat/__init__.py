import os
import re
from collections.abc import Iterable
from contextlib import contextmanager
from copy import copy
from typing import Any, Dict, List

import openpyxl
from jinja2.nativetypes import NativeEnvironment, NativeTemplate
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Cell, CellRange, Worksheet


class WombatSheet:
    new_worksheets: List["WombatSheet"] = []

    def __init__(self, title=None, context=None):
        self.title = title
        self.context = context or {}
        self.add_worksheet(self)

    @classmethod
    def add_worksheet(cls, ws: "WombatSheet"):
        cls.new_worksheets.append(ws)

    @classmethod
    @contextmanager
    def begin_adding_worksheets(cls):
        previous_worksheets = cls.new_worksheets
        cls.new_worksheets = []
        yield cls.new_worksheets
        cls.new_worksheets = previous_worksheets


class WombatEnvironment(NativeEnvironment):
    def __init__(self, *args, globals: Dict = None, **kwargs):
        super().__init__(*args, **kwargs)
        default_globals = {"WS": WombatSheet}
        if globals:
            default_globals.update(globals)
        self.globals.update(default_globals)


def render(workbook, environment):
    """
    Takes an openpyxl workbook and renders
    """
    render_worksheet_tabs(workbook, environment)

    for sheet in workbook:
        # TODO use worksheet environment
        render_worksheet(sheet, environment)

    return workbook


def render_worksheet(worksheet: Worksheet, env: WombatEnvironment) -> Worksheet:
    """
    Takes an openpyxl worksheet and renders it
    """
    for row_cells in worksheet.iter_rows():
        for cell in row_cells:
            value = cell.value

            if not value:
                continue

            output = env.from_string(value).render()
            if is_empty_rendered_value(output):
                cell.value = None
            else:
                # Shift cells right first before writing.
                if type(output) is not str and isinstance(output, Iterable):
                    output = list(output)
                    width = len(output)
                    worksheet.move_range(
                        cell_range=CellRange(
                            min_row=cell.row,
                            max_row=cell.row,
                            min_col=cell.column + 1,
                            max_col=100,
                        ),
                        cols=width - 1,
                        translate=True,
                    )

                    for idx, val in enumerate(output):
                        new_cell = worksheet.cell(cell.row, cell.column + idx)
                        new_cell.value = val
                        copy_style(original=cell, destination=new_cell)
                # Single cell template
                else:
                    cell.value = output
    return worksheet


def is_empty_rendered_value(output: Any):
    return output is None or (type(output) == str and output.strip() == "")


def is_for_loop(text: str) -> bool:
    return re.match("{%.*for.*%}.*{%.*endfor.*%}", text.strip(), re.MULTILINE)


def render_worksheet_tabs(workbook: Workbook, env: WombatEnvironment) -> Workbook:
    """
    Takes an workbook and templates the names
    """
    for worksheet in workbook.worksheets:
        with WombatSheet.begin_adding_worksheets():
            output = env.from_string(worksheet.title).render()

            if is_empty_rendered_value(output):
                workbook.remove(worksheet)
            elif WombatSheet.new_worksheets:
                offset = len(workbook.sheetnames) - workbook.index(worksheet)

                for ws in WombatSheet.new_worksheets:
                    new_worksheet = workbook.copy_worksheet(worksheet)
                    new_worksheet.title = str(ws.title)
                    # Adding wombat sheet to the OpenPyxl Worksheet
                    # So context of injected data (via jinja) is remembered.
                    new_worksheet.wombat_sheet = ws
                    offset = workbook.index(new_worksheet) - workbook.index(worksheet)
                    workbook.move_sheet(new_worksheet, -offset)

                workbook.remove(worksheet)
            else:
                worksheet.title = str(output).strip()
    return workbook


def copy_style(original: Cell, destination: Cell):
    """
    This method is using a private method... no other way
    for now thats as easy.
    """
    destination._style = copy(original._style)