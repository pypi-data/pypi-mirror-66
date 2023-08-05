from collections import OrderedDict

from openpyxl import load_workbook

from djaxei.exceptions import ImportException


class Importer:
    def __init__(self, tmpdir=None, **kwargs):
        self.tmpdir = tmpdir

    def xls_import(self, file, models_dict, *args, **kwargs):
        wb = load_workbook(file, data_only=True)

        for ws_name, callback in models_dict.items():
            ws = wb[ws_name]
            for rownum, row in enumerate(ws.iter_rows()):
                row = [cell.value for cell in row]
                if rownum == 0:
                    header = row
                else:
                    try:
                        data_dict = OrderedDict(zip(header, row))
                        data_dict = {k: v for k, v in data_dict.items() if k != None}
                        callback(data_dict)
                    except Exception as e:
                        raise ImportException(e, worksheet=ws_name)
        wb.close()

